import streamlit as st
import pandas as pd
import csv
import re
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Page configuration
st.set_page_config(
    page_title="Student Email Generator",
    page_icon="📧",
    layout="wide"
)

def parse_classroom_data(data):
    """
    Parse Google Classroom grade data and convert to CSV format.
    
    Args:
        data (str): Raw text data from Google Classroom
        
    Returns:
        str: CSV formatted string
    """
    # Split data into blocks
    blocks = data.split('\n\n')
    
    # Parse assignments
    course_block = blocks.pop(0)
    assignments = []
    lines = course_block.strip().split('\n')
    for idx, line in enumerate(lines):
        get = False
        if idx == 0:
            continue
        if ('assignment' in line.lower()) or ('quiz' in line.lower()) or ('form' in line.lower()) or ('project' in line.lower()) or ('book' in line.lower()):
            get = True
        if get:
            name = line
            next_line = lines[idx+1]
            if 'out of' in next_line:
                points = next_line.strip('out of ')
                assignments.append(f"{name} [{points}]")
            else:
                assignments.append(f"{name} [0]")
    
    # Parse student data
    students = []
    for block in blocks:
        lines = [l.strip() for l in block.strip().split('\n') if l.strip()]
        
        if not lines:
            continue
        # Student blocks start with a name
        student_name = lines[0]
        
        # Parse grades from remaining lines
        grades = []
        # Start at third element because 2nd one is overall grade (not necessary)
        i = 2
        while i < len(lines):
            line = lines[i]
            if 'Assigned' in line:
                grades.append('Pending')
                i += 1
                continue
            
            # Check for "/20No grade" pattern (submitted but not graded)
            if re.match(r'/\d+No grade', line):
                status = 'Submitted'
                next_line = line = lines[i+1]
                if 'Done late' in next_line:
                    status = 'Late'
                grades.append(status)
                i += 1
                continue
            
            # Check for "X out of Y" pattern
            match = re.search(r'^([\d.]+)\s+([\d.]+\s+)?out of [\d.]+', line)
            if match:
                grade = match.group(1)
                if grade != '0':
                    grades.append(grade)
                i += 1
                continue
            
            # Check for "0 out of X Draft•Missing" or similar
            if 'Missing' in line:
                grades.append('Missing')
                i += 1
                continue
            
            if 'Excused' in line:
                grades.append('Excused')
                i += 1
                continue

            if 'Turned in' in line:
                grades.append('Turned in')
                i += 1
                continue
            
            i += 1
        
        students.append([student_name] + grades)
    
    # Create CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    header = ['Student Name'] + assignments
    writer.writerow(header)
    
    # Write student rows
    for student in students:
        # Pad with empty strings if needed
        while len(student) < len(header):
            student.append('')
        writer.writerow(student[:len(header)])
    
    return output.getvalue()

def parse_assignment_name_and_points(assignment_string):
    """Extract assignment name and points from format 'Assignment Name [20]'"""
    match = re.match(r'^(.+?)\s*\[(\d+)\]$', assignment_string)
    if match:
        name = match.group(1).strip()
        points = int(match.group(2))
        return name, points
    return assignment_string, 10  # Default to 10 if parsing fails

def parse_grade(grade_value):
    """Parse grade value and return status and points."""
    if pd.isna(grade_value) or str(grade_value).strip() in ["", "-"]:
        return None, "Not yet assigned"
    
    grade_str = str(grade_value).strip()
    
    # Handle "Late" or "Late:" format
    if grade_str.lower().startswith("late"):
        if ":" in grade_str:
            try:
                points = float(grade_str.split(":")[1].strip())
                return points, "Done Late"
            except:
                return None, "Done Late"
        return None, "Done Late"
    
    if grade_str.lower() in ["missing", "not submitted"]:
        return 0, "Missing"
    
    if grade_str.lower() in ["submitted", "not graded yet", "ungraded"]:
        return None, "Submitted"

    if grade_str.lower() in ["pending"]:
        return None, "Pending"
    
    if grade_str.lower() == "excused":
        return None, "Excused"
    
    try:
        points = float(grade_str)
        return points, "Graded"
    except:
        return None, "Unknown"

def format_assignment_line(assignment_name, grade_value, max_points, is_assigned):
    """Format a single assignment line for the email."""
    if not is_assigned:
        return f"• {assignment_name}: worth {max_points} points"

    points, status = parse_grade(grade_value)

    # Format points as integer if it's a whole number
    if points is not None and points == int(points):
        points = int(points)
    if max_points == int(max_points):
        max_points = int(max_points)

    if status == "Graded":
        return f"• {assignment_name}: <b>{points}</b>/{max_points} points"
    elif status == "Done Late":
        if points is not None:
            return f"• {assignment_name}: <b>{points}</b>/{max_points} points (Done Late)"
        else:
            return f"• {assignment_name}: Done Late/{max_points} points"
    elif status == "Missing":
        return f"• {assignment_name}: <b>MISSING</b>/{max_points} points"
    elif status == "Submitted":
        return f"• {assignment_name}: Submitted/{max_points} points (Pending Grade)"
    elif status == "Pending":
        return f"• {assignment_name}: Not submitted/{max_points} points"
    elif status == "Excused":
        return f"• {assignment_name}: Excused"
    else:
        return f"• {assignment_name}: _/{max_points} points"

def calculate_total_points(row, assignments_config):
    """Calculate total points earned so far."""
    total = 0
    for assignment, config in assignments_config.items():
        if config["assigned"] and not config.get("omitted", False):
            grade_value = row.get(assignment)
            if grade_value is not None:
                points, status = parse_grade(grade_value)
                if points is not None and status in ["Graded", "Done Late"]:
                    total += points
    return int(total) if total == int(total) else total

def calculate_total_available_points(assignments_config):
    """Calculate total points available in the course (assigned + upcoming, excluding omitted)."""
    total = 0
    for assignment, config in assignments_config.items():
        if not config.get("omitted", False):
            total += config["max_points"]
    return int(total) if total == int(total) else total

def generate_email_body(row, assignments_config):
    """Generate personalized email body for a student."""
    total_points = calculate_total_points(row, assignments_config)
    total_available = calculate_total_available_points(assignments_config)
    
    progress_lines = []
    upcoming_lines = []
    
    for assignment, config in assignments_config.items():
        # Skip omitted assignments
        if config.get("omitted", False):
            continue
            
        grade_value = row.get(assignment)
        line = format_assignment_line(assignment, grade_value, config["max_points"], config["assigned"])
        
        if config["assigned"]:
            progress_lines.append(line)
        else:
            upcoming_lines.append(line)
    
    progress_section = "\n".join(progress_lines)
    upcoming_section = "\n".join(upcoming_lines)
    
    email_body = f"""<b>CURRENT TOTAL: {total_points} points</b>

✍️ <b>Progress so far:</b>
{progress_section}

📖 <b>Upcoming Assignments:</b>
{upcoming_section}

<b>TOTAL POINTS AVAILABLE: {total_available}</b>"""
    
    return email_body

def generate_excel_with_formatting(df, assignments_config):
    """Generate Excel file with color coding and analytics."""
    wb = Workbook()
    
    # Define colors
    light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Sheet 1: Student Data
    ws1 = wb.active
    ws1.title = "Student Data"
    
    # Calculate Total Points Earned and Certificate Status
    total_points_list = []
    certificate_status_list = []
    
    for idx, row in df.iterrows():
        total_points = calculate_total_points(row, assignments_config)
        total_points_list.append(total_points)
        
        if total_points < 40:
            status = "None"
        elif total_points < 80:
            status = "Participation"
        else:
            status = "Completion"
        certificate_status_list.append(status)
    
    # Filter out omitted assignment columns
    omitted_assignments = [name for name, config in assignments_config.items() if config.get("omitted", False)]
    df_export = df.drop(columns=omitted_assignments, errors='ignore').copy()
    
    # Add calculated columns to filtered dataframe
    df_export.insert(1, 'Total Points Earned', total_points_list)
    df_export.insert(2, 'Certificate Status', certificate_status_list)
    
    # Write data to sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            # Format header row
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Color code Total Points Earned (column B)
                if c_idx == 2:
                    if isinstance(value, (int, float)):
                        if value >= 80:
                            cell.fill = light_green
                        elif value >= 40:
                            cell.fill = light_yellow
                
                # Color code Certificate Status (column C)
                elif c_idx == 3:
                    if value == "Completion":
                        cell.fill = light_green
                    elif value == "Participation":
                        cell.fill = light_yellow
                
                # Color code assignment columns for Submitted/Pending
                elif c_idx > 3:
                    if isinstance(value, str):
                        value_lower = value.lower()
                        if any(keyword in value_lower for keyword in ['submitted', 'pending', 'not graded']):
                            cell.fill = light_yellow
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Class Analytics
    ws2 = wb.create_sheet("Class Analytics")
    
    # Calculate analytics data
    total_students = len(df_export)
    class_avg = sum(total_points_list) / total_students if total_students > 0 else 0
    class_median = pd.Series(total_points_list).median()
    
    none_count = certificate_status_list.count('None')
    participation_count = certificate_status_list.count('Participation')
    completion_count = certificate_status_list.count('Completion')
    
    # Write analytics data
    current_row = 1
    
    # Student Performance section
    ws2.cell(row=current_row, column=1, value="STUDENT PERFORMANCE").font = Font(bold=True, size=14)
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Total Students")
    ws2.cell(row=current_row, column=2, value=total_students)
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Class Average")
    ws2.cell(row=current_row, column=2, value=round(class_avg, 2))
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Class Median")
    ws2.cell(row=current_row, column=2, value=class_median)
    current_row += 2
    
    # Certificate Distribution section
    cert_dist_start = current_row
    ws2.cell(row=current_row, column=1, value="CERTIFICATE DISTRIBUTION").font = Font(bold=True, size=14)
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Status")
    ws2.cell(row=current_row, column=2, value="Count")
    ws2.cell(row=current_row, column=3, value="Percentage")
    current_row += 1
    
    cert_data_start = current_row
    ws2.cell(row=current_row, column=1, value="None")
    ws2.cell(row=current_row, column=2, value=none_count)
    ws2.cell(row=current_row, column=3, value=f"{(none_count/total_students*100):.1f}%")
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Participation")
    ws2.cell(row=current_row, column=2, value=participation_count)
    ws2.cell(row=current_row, column=3, value=f"{(participation_count/total_students*100):.1f}%")
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Completion")
    ws2.cell(row=current_row, column=2, value=completion_count)
    ws2.cell(row=current_row, column=3, value=f"{(completion_count/total_students*100):.1f}%")
    current_row += 2
    
    # Grade Distribution section
    grade_dist_start = current_row
    ws2.cell(row=current_row, column=1, value="GRADE DISTRIBUTION").font = Font(bold=True, size=14)
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Range")
    ws2.cell(row=current_row, column=2, value="Count")
    current_row += 1
    
    ranges = [(0, 19), (20, 39), (40, 59), (60, 79), (80, 99), (100, 1000)]
    range_labels = ['0-19', '20-39', '40-59', '60-79', '80-99', '100+']
    
    grade_data_start = current_row
    for label, (low, high) in zip(range_labels, ranges):
        count = sum(1 for pts in total_points_list if low <= pts <= high)
        ws2.cell(row=current_row, column=1, value=label)
        ws2.cell(row=current_row, column=2, value=count)
        current_row += 1
    grade_data_end = current_row - 1
    current_row += 1
    
    # Pending Assignments section
    ws2.cell(row=current_row, column=1, value="PENDING ASSIGNMENTS").font = Font(bold=True, size=14)
    current_row += 1
    total_pending = 0
    for idx, row_data in df.iterrows():
        for assignment, config in assignments_config.items():
            if not config.get("omitted", False):
                grade_value = row_data.get(assignment)
                points, status = parse_grade(grade_value)
                if status == "Submitted":
                    total_pending += 1
    ws2.cell(row=current_row, column=1, value="Total Pending Assignments")
    ws2.cell(row=current_row, column=2, value=total_pending)
    current_row += 2
    
    # Assignment Completion Rates section
    ws2.cell(row=current_row, column=1, value="ASSIGNMENT COMPLETION RATES").font = Font(bold=True, size=14)
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Assignment Name")
    ws2.cell(row=current_row, column=2, value="Completion Rate")
    current_row += 1
    
    for assignment, config in assignments_config.items():
        if not config.get("omitted", False):
            completed_count = 0
            for idx, row_data in df.iterrows():
                grade_value = row_data.get(assignment)
                points, status = parse_grade(grade_value)
                if status in ["Graded", "Submitted", "Done Late"]:
                    completed_count += 1
            completion_rate = (completed_count / total_students * 100) if total_students > 0 else 0
            ws2.cell(row=current_row, column=1, value=assignment)
            ws2.cell(row=current_row, column=2, value=f"{completion_rate:.1f}%")
            current_row += 1
    current_row += 1
    
    # Top 10% Students section
    ws2.cell(row=current_row, column=1, value="TOP 10% STUDENTS").font = Font(bold=True, size=14)
    current_row += 1
    ws2.cell(row=current_row, column=1, value="Student Name")
    ws2.cell(row=current_row, column=2, value="Total Points")
    current_row += 1
    
    student_scores = [(df_export.iloc[i]['Student Name'], total_points_list[i]) for i in range(len(df_export))]
    student_scores.sort(key=lambda x: x[1], reverse=True)
    
    top_10_percent_count = max(1, int(total_students * 0.1))
    if top_10_percent_count < len(student_scores):
        threshold_score = student_scores[top_10_percent_count - 1][1]
        top_students = [s for s in student_scores if s[1] >= threshold_score]
    else:
        top_students = student_scores
    
    for student_name, points in top_students:
        ws2.cell(row=current_row, column=1, value=student_name)
        ws2.cell(row=current_row, column=2, value=points)
        current_row += 1
    
    # Add Certificate Distribution Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Certificate Distribution"
    pie_chart.style = 10
    
    labels = Reference(ws2, min_col=1, min_row=cert_data_start, max_row=cert_data_start + 2)
    data = Reference(ws2, min_col=2, min_row=cert_data_start - 1, max_row=cert_data_start + 2)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    
    ws2.add_chart(pie_chart, "E6")
    
    # Add Grade Distribution Bar Chart
    bar_chart = BarChart()
    bar_chart.title = "Grade Distribution"
    bar_chart.x_axis.title = "Grade Range"
    bar_chart.y_axis.title = "Number of Students"
    bar_chart.style = 10
    
    labels = Reference(ws2, min_col=1, min_row=grade_data_start, max_row=grade_data_end)
    data = Reference(ws2, min_col=2, min_row=grade_data_start - 1, max_row=grade_data_end)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(labels)
    
    ws2.add_chart(bar_chart, "E22")
    
    # Auto-adjust column widths for analytics sheet
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

# Initialize session state
if 'sent_status' not in st.session_state:
    st.session_state.sent_status = {}
if 'generated_data' not in st.session_state:
    st.session_state.generated_data = None
if 'assignments_config' not in st.session_state:
    st.session_state.assignments_config = {}
if 'current_df' not in st.session_state:
    st.session_state.current_df = None
if 'raw_data' not in st.session_state:
    st.session_state.raw_data = ""
if 'cleaned_csv' not in st.session_state:
    st.session_state.cleaned_csv = None
if 'show_add_assignment' not in st.session_state:
    st.session_state.show_add_assignment = False

# Custom CSS
st.markdown("""
<style>
    .student-card {
        background-color: #f0f2f6;
        padding: 20px;
        border-radius: 10px;
        margin-bottom: 20px;
    }
    .student-name {
        font-size: 24px;
        font-weight: bold;
        color: #1f77b4;
    }
    .points-badge {
        font-size: 18px;
        font-weight: bold;
        padding: 5px 10px;
        border-radius: 5px;
        display: inline-block;
    }
    .points-high {
        background-color: #d4edda;
        color: #155724;
    }
    .points-medium {
        background-color: #fff3cd;
        color: #856404;
    }
    .points-low {
        background-color: #f8d7da;
        color: #721c24;
    }
</style>
""", unsafe_allow_html=True)

# Streamlit UI
st.title("📧 Student Grade Summary Generator")
st.markdown("### Quick Grade Snippets for HubSpot")

# Sidebar for configuration
with st.sidebar:
    st.header("⚙️ Configuration")
    
    if st.session_state.current_df is not None and st.session_state.assignments_config:
        st.markdown("### 📝 Assignment Settings")
        st.markdown("*Adjust points and mark as Assigned/Upcoming*")
        
        with st.expander("✏️ Configure Assignments", expanded=True):
            # Show all assignments
            assignment_keys = list(st.session_state.assignments_config.keys())
            
            for idx, name in enumerate(assignment_keys):
                config = st.session_state.assignments_config[name]
                st.markdown(f"**{name}**")
                col1, col2 = st.columns([3, 2])
                
                with col1:
                    new_assigned = st.checkbox(
                        "✓ Assigned",
                        value=config["assigned"] and not config.get("omitted", False),
                        key=f"assigned_{idx}_{name}",
                        help="Check if this assignment has been assigned to students",
                        disabled=config.get("omitted", False)
                    )
                
                with col2:
                    new_points = st.number_input(
                        "Max Points",
                        value=config["max_points"],
                        min_value=0,
                        step=1,
                        key=f"points_{idx}_{name}",
                        help="Maximum points for this assignment"
                    )
                
                # Add Omit checkbox below
                new_omitted = st.checkbox(
                    "⊘ Omit from emails",
                    value=config.get("omitted", False),
                    key=f"omitted_{idx}_{name}",
                    help="Check to exclude this assignment from email summaries"
                )
                
                # If omitted, automatically uncheck assigned
                if new_omitted:
                    new_assigned = False
                
                st.session_state.assignments_config[name] = {
                    "max_points": new_points,
                    "assigned": new_assigned,
                    "omitted": new_omitted
                }
                
                st.markdown("---")
        
        # Add new assignment section (inline in sidebar)
        st.markdown("### ➕ Add New Assignment")
        if st.session_state.show_add_assignment:
            new_assignment_name = st.text_input("Assignment Name", placeholder="e.g., Final Project", key="new_assign_name")
            new_assignment_points = st.number_input("Max Points", min_value=1, value=10, step=1, key="new_assign_points")
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Add", type="primary", use_container_width=True):
                    if new_assignment_name:
                        # Add to assignments config as upcoming
                        st.session_state.assignments_config[new_assignment_name] = {
                            "max_points": new_assignment_points,
                            "assigned": False,
                            "omitted": False
                        }
                        
                        # Add empty column to dataframe if it exists
                        if st.session_state.current_df is not None:
                            st.session_state.current_df[new_assignment_name] = ""
                        
                        st.session_state.show_add_assignment = False
                        st.success(f"✅ Added '{new_assignment_name}'")
                        st.rerun()
            with col2:
                if st.button("Cancel", use_container_width=True):
                    st.session_state.show_add_assignment = False
                    st.rerun()
        else:
            if st.button("➕ Add New Assignment", use_container_width=True):
                st.session_state.show_add_assignment = True
                st.rerun()
        
        # Show current configuration summary
        with st.expander("👁️ View Current Setup"):
            assigned_assignments = {k: v for k, v in st.session_state.assignments_config.items() if v["assigned"] and not v.get("omitted", False)}
            upcoming_assignments = {k: v for k, v in st.session_state.assignments_config.items() if not v["assigned"] and not v.get("omitted", False)}
            omitted_assignments = {k: v for k, v in st.session_state.assignments_config.items() if v.get("omitted", False)}
            
            st.markdown("**Assigned Assignments:**")
            if assigned_assignments:
                for name, config in assigned_assignments.items():
                    st.write(f"• {name}: {config['max_points']} points")
            else:
                st.write("_(None marked as assigned yet)_")
            
            st.markdown("**Upcoming Assignments:**")
            if upcoming_assignments:
                for name, config in upcoming_assignments.items():
                    st.write(f"• {name}: {config['max_points']} points")
            else:
                st.write("_(All assignments marked as assigned)_")
            
            st.markdown("**⊘ Omitted Assignments:**")
            if omitted_assignments:
                for name, config in omitted_assignments.items():
                    st.write(f"• {name}: {config['max_points']} points")
            else:
                st.write("_(No assignments omitted)_")
            
            total_assigned_points = sum(c["max_points"] for c in assigned_assignments.values())
            total_all_points = sum(c["max_points"] for c in st.session_state.assignments_config.values() if not c.get("omitted", False))
            st.info(f"📊 Total Assigned: {total_assigned_points} | Total Course: {total_all_points}")
    else:
        st.info("📋 Paste Google Classroom data to begin")
    
    st.markdown("---")
    if st.session_state.generated_data is not None:
        sent_count = sum(1 for v in st.session_state.sent_status.values() if v)
        total_count = len(st.session_state.generated_data)
        st.metric("📊 Progress", f"{sent_count}/{total_count} sent")
        
        if sent_count > 0:
            progress_pct = (sent_count / total_count) * 100
            st.progress(progress_pct / 100)

# Main content area - Data input
st.header("📋 Step 1: Paste Google Classroom Data")

raw_text = st.text_area(
    "Paste raw Google Classroom grade data here",
    height=300,
    placeholder="Paste your Google Classroom data here...",
    help="Copy and paste the grade data directly from Google Classroom"
)

# Auto-process when text is pasted
if raw_text and raw_text != st.session_state.raw_data:
    st.session_state.raw_data = raw_text
    
    try:
        with st.spinner("🔄 Cleaning data..."):
            # Clean the data
            cleaned_csv = parse_classroom_data(raw_text)
            st.session_state.cleaned_csv = cleaned_csv
            
            # Parse CSV into dataframe
            df = pd.read_csv(StringIO(cleaned_csv))
            
            # Rename columns to remove brackets and extract assignment config
            column_mapping = {}
            st.session_state.assignments_config = {}
            
            for col in df.columns:
                if col != "Student Name":
                    name, points = parse_assignment_name_and_points(col)
                    column_mapping[col] = name  # Map old name to new clean name
                    # Check if assignment has any grades
                    has_grades = df[col].notna().any() and not all(df[col].fillna("").astype(str).str.strip() == "")
                    st.session_state.assignments_config[name] = {
                        "max_points": points,
                        "assigned": has_grades,
                        "omitted": False
                    }
            
            # Rename DataFrame columns to clean names (without brackets)
            df = df.rename(columns=column_mapping)
            st.session_state.current_df = df
            
            # Reset generated data when new data is pasted
            st.session_state.generated_data = None
            st.session_state.sent_status = {}
            
        st.success("✅ Data cleaned successfully!")
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Error processing data: {str(e)}")
        st.info("Please make sure you've pasted valid Google Classroom data.")

# Step 2: Configure Assignments (reference to sidebar)
if st.session_state.current_df is not None:
    st.markdown("---")
    st.header("⚙️ Step 2: Configure Assignments")
    st.info("👈 Use the sidebar to configure which assignments are 'Assigned' vs 'Upcoming', adjust point values, and add new assignments.")

# Show cleaned CSV preview - Step 3
if st.session_state.current_df is not None and st.session_state.assignments_config:
    st.markdown("---")
    st.header("📊 Step 3: Review Cleaned Student Data")
    
    with st.expander("🔍 Preview Cleaned CSV", expanded=True):
        st.dataframe(st.session_state.current_df)
    
    # Download button for Excel file with formatting
    st.markdown("### 📥 Download Data")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Generate Excel file
        excel_data = generate_excel_with_formatting(st.session_state.current_df, st.session_state.assignments_config)
        
        st.download_button(
            label="📊 Download Excel Report (with colors & charts)",
            data=excel_data,
            file_name="classroom_data_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Downloads formatted Excel file with color coding and analytics charts"
        )
    
    with col2:
        # Also keep CSV option for compatibility
        download_df = st.session_state.current_df.copy()
        
        total_points_list = []
        certificate_status_list = []
        
        for idx, row in download_df.iterrows():
            total_points = calculate_total_points(row, st.session_state.assignments_config)
            total_points_list.append(total_points)
            
            if total_points < 40:
                status = "None"
            elif total_points < 80:
                status = "Participation"
            else:
                status = "Completion"
            certificate_status_list.append(status)
        
        download_df.insert(1, 'Total Points Earned', total_points_list)
        download_df.insert(2, 'Certificate Status', certificate_status_list)
        
        csv_buffer = StringIO()
        download_df.to_csv(csv_buffer, index=False)
        
        st.download_button(
            label="⬇️ Download CSV (basic)",
            data=csv_buffer.getvalue(),
            file_name="cleaned_classroom_data.csv",
            mime="text/csv",
            use_container_width=True,
            help="Downloads basic CSV without formatting"
        )
    
    st.markdown("---")
    st.header("📧 Step 4: Generate Student Email Snippets")
    
    # Generate emails button
    if st.button("🚀 Generate Grade Summaries", type="primary", use_container_width=True):
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        df = st.session_state.current_df
        
        for idx, row in df.iterrows():
            grade_summary = generate_email_body(row, st.session_state.assignments_config)
            total_points = calculate_total_points(row, st.session_state.assignments_config)
            
            student_id = f"{row['Student Name']}"
            
            results.append({
                "student_id": student_id,
                "Student Name": row["Student Name"],
                "Grade Summary": grade_summary,
                "Total Points": total_points
            })
            
            if student_id not in st.session_state.sent_status:
                st.session_state.sent_status[student_id] = False
            
            progress_bar.progress((idx + 1) / len(df))
            status_text.text(f"Processing {idx + 1}/{len(df)}: {row['Student Name']}")
        
        status_text.text("✅ All emails generated!")
        st.session_state.generated_data = results
        
        # Display statistics
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        output_df = pd.DataFrame(results)
        with col1:
            st.metric("Total Students", len(results))
        with col2:
            avg_points = output_df["Total Points"].mean()
            st.metric("Average Points", f"{avg_points:.1f}")
        with col3:
            completion_eligible = len(output_df[output_df["Total Points"] >= 80])
            st.metric("On Track for Completion", completion_eligible)

# Display individual student grade summaries
if st.session_state.generated_data is not None:
    st.markdown("---")
    st.header("📨 Individual Student Grade Summaries")
    
    # Filter options
    col1, col2 = st.columns(2)
    with col1:
        show_sent = st.checkbox("Show already sent", value=True)
    with col2:
        show_unsent = st.checkbox("Show not sent", value=True)
    
    for idx, student_data in enumerate(st.session_state.generated_data):
        student_id = student_data["student_id"]
        is_sent = st.session_state.sent_status.get(student_id, False)
        
        if (is_sent and not show_sent) or (not is_sent and not show_unsent):
            continue
        
        points = student_data["Total Points"]
        if points >= 80:
            badge_class = "points-high"
            badge_text = "🏆 Completion Track"
        elif points >= 40:
            badge_class = "points-medium"
            badge_text = "📜 Participation Track"
        else:
            badge_class = "points-low"
            badge_text = "⚠️ Below Threshold"
        
        with st.container():
            st.markdown(f"""
            <div class="student-card">
                <span class="student-name">
                    {student_data['Student Name']}
                </span>
                <span class="points-badge {badge_class}">
                    {points} points - {badge_text}
                </span>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.text(f"👤 {student_data['Student Name']}")
            
            with col2:
                sent = st.checkbox(
                    "✓ Sent" if is_sent else "Mark Sent",
                    value=is_sent,
                    key=f"sent_{student_id}"
                )
                st.session_state.sent_status[student_id] = sent
            
            with st.expander("👁️ Preview Grade Summary"):
                # Formatted HTML version for copying
                st.caption("📋 **Formatted Version (Select and copy this into HubSpot):**")
                st.markdown(
                    f"""<div style="padding: 15px; border-radius: 5px; border: 1px solid #ddd; font-family: Arial, sans-serif; user-select: all;">
                    {student_data["Grade Summary"].replace(chr(10), '<br>')}
                    </div>""",
                    unsafe_allow_html=True
                )
                st.info("💡 Click inside the box above, press Ctrl+A (or Cmd+A) to select all, then Ctrl+C (or Cmd+C) to copy.")
            
            st.markdown("---")
    
    # Download options
    st.markdown("### 💾 Export Options")
    col1, col2 = st.columns(2)
    
    with col1:
        output_df = pd.DataFrame(st.session_state.generated_data)
        output_df = output_df.drop('student_id', axis=1)
        csv_buffer = StringIO()
        output_df.to_csv(csv_buffer, index=False, quoting=csv.QUOTE_ALL)
        csv_string = csv_buffer.getvalue()
        
        st.download_button(
            label="⬇️ Download All Emails (CSV)",
            data=csv_string,
            file_name="hubspot_import.csv",
            mime="text/csv"
        )
    
    with col2:
        sent_log = []
        for student in st.session_state.generated_data:
            sent_log.append({
                "Name": student['Student Name'],
                "Sent": "Yes" if st.session_state.sent_status.get(student['student_id'], False) else "No"
            })
        
        log_df = pd.DataFrame(sent_log)
        log_buffer = StringIO()
        log_df.to_csv(log_buffer, index=False)
        
        st.download_button(
            label="📊 Download Send Log",
            data=log_buffer.getvalue(),
            file_name="email_send_log.csv",
            mime="text/csv"
        )

# Footer
st.markdown("---")
st.markdown("Made for Google Classroom | Auto-cleans and processes grade data")
